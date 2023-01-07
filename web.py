from flask import Flask, render_template, request, make_response
from openpyxl import load_workbook
from partidos_quiniela import partidos

app = Flask(__name__)


@app.route("/")
def inicio():
    return render_template("inicio.html")


@app.route("/predicciones")
def predicciones():
    return render_template("predicciones.html")


@app.route("/resultados")
def resultados():
    book = load_workbook("Resultados.xlsx")
    sheet = book.active
    return render_template("resultados.html", sheet=sheet)


@app.route("/predicciones/jornada-18")
def pred_j18():
    book = load_workbook("datos.xlsx")
    sheet = book.active
    partidos_list = partidos(18)
    jornada = 1
    book2 = load_workbook("Resultados.xlsx")
    resultados = book2.active
    return render_template("18.html", sheet=sheet, partidos_list=partidos_list, jornada=jornada, resultados=resultados)

@app.route("/predicciones/jornada-20")
def pred_j20():
    book = load_workbook("datos.xlsx")
    sheet = book.active
    partidos_list = partidos(20)
    jornada = 2
    book2 = load_workbook("Resultados.xlsx")
    resultados = book2.active
    return render_template("20.html", sheet=sheet, partidos_list=partidos_list, jornada=jornada, resultados=resultados)

@app.route("/predicciones/jornada-26")
def pred_j26():
    book = load_workbook("datos.xlsx")
    sheet = book.active
    partidos_list = partidos(26)
    jornada = 3
    book2 = load_workbook("Resultados.xlsx")
    resultados = book2.active
    return render_template("26.html", sheet=sheet, partidos_list=partidos_list, jornada=jornada, resultados=resultados)

@app.route("/predicciones/jornada-27")
def pred_j27():
    book = load_workbook("datos.xlsx")
    sheet = book.active
    partidos_list = partidos(27)
    jornada = 4
    book2 = load_workbook("Resultados.xlsx")
    resultados = book2.active
    return render_template("27.html", sheet=sheet, partidos_list=partidos_list, jornada=jornada, resultados=resultados)

@app.route("/predicciones/jornada-32")
def pred_j32():
    book = load_workbook("datos.xlsx")
    sheet = book.active
    partidos_list = partidos(32)
    jornada = 5
    book2 = load_workbook("Resultados.xlsx")
    resultados = book2.active
    return render_template("32.html", sheet=sheet, partidos_list=partidos_list, jornada=jornada, resultados=resultados)

if __name__ == "__main__":
    app.run(debug=True)
